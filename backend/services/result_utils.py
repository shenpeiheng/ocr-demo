"""
OCR/PDF 结果处理公共函数。
"""

import base64
import io
import json
import os

import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image, ImageDraw, ImageFont


def draw_ocr_boxes(image_path, text_items):
    """在图片上绘制 OCR 框和序号标签。"""
    try:
        pil_img = Image.open(image_path).convert("RGB")
        img_width, img_height = pil_img.size
        img_cv = cv2.cvtColor(np.array(pil_img), cv2.COLOR_RGB2BGR)

        script_dir = os.path.dirname(os.path.abspath(__file__))
        backend_dir = os.path.dirname(script_dir)
        project_root = os.path.dirname(backend_dir)
        font_paths = [
            os.path.join(project_root, "frontend", "static", "fonts", "NotoSansSC.ttf"),
            "C:/Windows/Fonts/msyh.ttc",
            "C:/Windows/Fonts/simhei.ttf",
            "C:/Windows/Fonts/simsun.ttc",
            "C:/Windows/Fonts/yahei.ttf",
            "C:/Windows/Fonts/msyhbd.ttc",
        ]
        font_large = None
        font_small = None
        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    font_large = ImageFont.truetype(font_path, 20)
                    font_small = ImageFont.truetype(font_path, 16)
                    break
                except Exception:
                    continue
        if font_large is None:
            font_large = ImageFont.load_default()
            font_small = ImageFont.load_default()

        box_color_bgr = (0, 0, 255)
        label_bg_bgr = (255, 255, 255)
        label_text_color = (0, 0, 0)

        for index, item in enumerate(text_items):
            location = item.get("location", {})
            left = int(location.get("left", 0))
            top = int(location.get("top", 0))
            width = int(location.get("width", 50))
            height = int(location.get("height", 20))

            left = max(0, min(left, img_width - 1))
            top = max(0, min(top, img_height - 1))
            right = min(left + width, img_width)
            bottom = min(top + height, img_height)

            if right <= left or bottom <= top:
                continue

            cv2.rectangle(img_cv, (left, top), (right, bottom), box_color_bgr, 2)

            label = f"#{index + 1}"
            pil_tmp = Image.fromarray(cv2.cvtColor(img_cv, cv2.COLOR_BGR2RGB))
            draw_tmp = ImageDraw.Draw(pil_tmp)
            bbox = draw_tmp.textbbox((0, 0), label, font=font_small)
            label_w = bbox[2] - bbox[0] + 12
            label_h = bbox[3] - bbox[1] + 6

            label_x = left
            label_y = top - label_h - 2
            if label_y < 0:
                label_y = top + 2

            cv2.rectangle(
                img_cv,
                (label_x, label_y),
                (label_x + label_w, label_y + label_h),
                label_bg_bgr,
                -1,
            )
            cv2.rectangle(
                img_cv,
                (label_x, label_y),
                (label_x + label_w, label_y + label_h),
                box_color_bgr,
                1,
            )

            pil_tmp = Image.fromarray(cv2.cvtColor(img_cv, cv2.COLOR_BGR2RGB))
            draw_tmp = ImageDraw.Draw(pil_tmp)
            draw_tmp.text(
                (label_x + 6, label_y + 3),
                label,
                fill=label_text_color,
                font=font_small,
            )
            img_cv = cv2.cvtColor(np.array(pil_tmp), cv2.COLOR_RGB2BGR)

        final_img = Image.fromarray(cv2.cvtColor(img_cv, cv2.COLOR_BGR2RGB))
        output_buffer = io.BytesIO()
        final_img.save(output_buffer, format="JPEG", quality=95)
        img_base64 = base64.b64encode(output_buffer.getvalue()).decode("utf-8")
        return f"data:image/jpeg;base64,{img_base64}"
    except Exception as exc:
        print(f"[draw_ocr_boxes] 错误: {exc}")
        import traceback

        traceback.print_exc()
        return None


def format_results_as_json(results):
    """将 OCR 结果格式化为结构化 JSON 字符串。"""
    try:
        text_items = results.get("text_items", [])
        image_info = results.get("image_info", {})
        analysis = results.get("analysis", {})

        formatted = {
            "识别统计": {
                "总识别项数": results.get("total_items", 0),
                "处理时间(秒)": results.get("processing_time", 0),
                "OCR引擎": results.get("ocr_engine", "unknown"),
                "图片宽度": image_info.get("width", 0),
                "图片高度": image_info.get("height", 0),
            },
            "类型分布": analysis.get("type_distribution", {}),
            "识别详情": [],
        }

        for index, item in enumerate(text_items):
            location = item.get("location", {})
            formatted["识别详情"].append(
                {
                    "序号": index + 1,
                    "内容": item.get("text", ""),
                    "类型": item.get("type", "text"),
                    "区域": item.get("region", ""),
                    "坐标": {
                        "left": location.get("left", 0),
                        "top": location.get("top", 0),
                        "width": location.get("width", 0),
                        "height": location.get("height", 0),
                    },
                    "置信度": item.get("confidence", 0),
                }
            )

        return json.dumps(formatted, ensure_ascii=False, indent=2)
    except Exception as exc:
        print(f"[format_results_as_json] 错误: {exc}")
        return json.dumps({"error": str(exc)}, ensure_ascii=False, indent=2)


def generate_excel(results, excel_path):
    """生成图片 OCR 结果 Excel。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "OCR识别结果"

    headers = ["序号", "坐标X", "坐标Y", "宽度", "高度", "内容", "置信度", "类型"]
    if "region" in results.get("text_items", [{}])[0]:
        headers.append("区域")
    ws.append(headers)

    for index, item in enumerate(results.get("text_items", []), 1):
        location = item.get("location", {})
        row_data = [
            index,
            location.get("left", 0),
            location.get("top", 0),
            location.get("width", 0),
            location.get("height", 0),
            item.get("text", ""),
            item.get("confidence", 0),
            item.get("type", "text"),
        ]
        if "region" in item:
            row_data.append(item.get("region", ""))
        ws.append(row_data)

    _adjust_sheet_widths(wb)
    wb.save(excel_path)


def generate_pdf_excel(pdf_result, excel_path):
    """生成 PDF OCR 结果 Excel。"""
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "PDF处理汇总"

    pdf_info = pdf_result.get("pdf_info", {})
    processing_info = pdf_result.get("processing_info", {})
    combined_results = pdf_result.get("combined_results", {})

    ws_summary.append(["PDF文件信息", ""])
    ws_summary.append(["文件名", pdf_info.get("filename", "")])
    ws_summary.append(["总页数", pdf_info.get("total_pages", 0)])
    ws_summary.append(["文件大小", f"{pdf_info.get('file_size', 0)} 字节"])
    ws_summary.append([])
    ws_summary.append(["处理信息", ""])
    ws_summary.append(["DPI", processing_info.get("dpi", 200)])
    ws_summary.append(["最大处理页数", processing_info.get("max_pages", 10)])
    ws_summary.append(["实际处理页数", processing_info.get("actual_pages_processed", 0)])
    ws_summary.append([])
    ws_summary.append(["识别结果汇总", ""])
    ws_summary.append(["总识别项数", combined_results.get("total_items", 0)])
    ws_summary.append(["总处理时间", f"{combined_results.get('total_processing_time', 0)} 秒"])

    ws_pages = wb.create_sheet(title="页面汇总")
    ws_pages.append(["页码", "识别项数", "处理时间(秒)", "处理状态"])
    for page in pdf_result.get("pages_summary", []):
        ws_pages.append(
            [
                page.get("page_number", 0),
                page.get("total_items", 0),
                page.get("processing_time", 0),
                "成功" if page.get("success", False) else "失败",
            ]
        )

    ws_details = wb.create_sheet(title="详细识别结果")
    detail_headers = ["页码", "序号", "坐标X", "坐标Y", "宽度", "高度", "内容", "置信度", "类型"]
    if "region" in combined_results.get("text_items", [{}])[0]:
        detail_headers.append("区域")
    ws_details.append(detail_headers)

    for index, item in enumerate(combined_results.get("text_items", []), 1):
        location = item.get("location", {})
        row_data = [
            item.get("page", 1),
            index,
            location.get("left", 0),
            location.get("top", 0),
            location.get("width", 0),
            location.get("height", 0),
            item.get("text", ""),
            item.get("confidence", 0),
            item.get("type", "text"),
        ]
        if "region" in item:
            row_data.append(item.get("region", ""))
        ws_details.append(row_data)

    _adjust_sheet_widths(wb)
    wb.save(excel_path)


def generate_flowchart_excel(flowchart_result, excel_path):
    """生成流程图识别结果 Excel。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "流程图识别结果"

    headers = ["流程", "流程说明", "流程ID", "流程描述", "操作方式", "部门"]
    ws.append(headers)

    header_gray = PatternFill("solid", fgColor="BFBFBF")
    header_blue = PatternFill("solid", fgColor="1F4E79")
    light_group_fill = PatternFill("solid", fgColor="DDEBFF")
    thin_side = Side(style="thin", color="000000")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_font = Font(bold=True, color="000000")
    header_white_font = Font(bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=10)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_blue if header in {"操作方式", "部门"} else header_gray
        cell.font = header_white_font if header in {"操作方式", "部门"} else header_font
        cell.alignment = center_alignment
        cell.border = border

    rows = flowchart_result.get("rows", [])
    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.font = body_font
            cell.alignment = left_alignment if cell.column == 4 else center_alignment
            cell.border = border
        row[0].fill = light_group_fill
        row[1].fill = light_group_fill

    _merge_flowchart_groups(ws)

    widths = {
        "A": 14,
        "B": 22,
        "C": 18,
        "D": 62,
        "E": 22,
        "F": 14,
    }
    for column_letter, width in widths.items():
        ws.column_dimensions[column_letter].width = width

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 26
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 22

    wb.save(excel_path)


def generate_word_flowchart_excel(word_result, excel_path):
    """生成 Word 流程图批量识别 Excel。"""
    wb = Workbook()
    thin_side = Side(style="thin", color="000000")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_gray = PatternFill("solid", fgColor="BFBFBF")
    header_blue = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="000000")
    header_white_font = Font(bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=10)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws = wb.active
    ws.title = "流程图识别结果"
    result_headers = ["来源文档", "图片序号", "来源图片", "流程", "流程说明", "流程ID", "流程描述", "操作方式", "部门"]
    ws.append(result_headers)
    _style_header_row(ws, result_headers, header_gray, header_blue, header_font, header_white_font, center_alignment, border)

    for row in word_result.get("rows", []):
        ws.append([row.get(header, "") for header in result_headers])

    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(result_headers)):
        for cell in row_cells:
            cell.font = body_font
            cell.alignment = left_alignment if cell.column == 7 else center_alignment
            cell.border = border

    ws.freeze_panes = "A2"
    result_widths = {
        "A": 28,
        "B": 10,
        "C": 18,
        "D": 14,
        "E": 22,
        "F": 18,
        "G": 62,
        "H": 22,
        "I": 14,
    }
    for column_letter, width in result_widths.items():
        ws.column_dimensions[column_letter].width = width

    ws_status = wb.create_sheet(title="图片处理清单")
    status_headers = ["图片序号", "来源图片", "处理状态", "流程节点数", "处理时间(秒)", "失败原因"]
    ws_status.append(status_headers)
    _style_header_row(
        ws_status,
        status_headers,
        header_gray,
        header_blue,
        header_font,
        header_white_font,
        center_alignment,
        border,
        blue_headers={"处理状态", "失败原因"},
    )

    for file_result in word_result.get("files", []):
        status = file_result.get("status", "")
        success = bool(file_result.get("success"))
        skipped = status == "skipped" or file_result.get("selected") is False
        status_label = "未选择" if skipped else ("成功" if success else "失败")
        ws_status.append(
            [
                file_result.get("image_index", ""),
                file_result.get("original_filename", file_result.get("filename", "")),
                status_label,
                file_result.get("total_rows", 0),
                file_result.get("processing_time", ""),
                "" if success or skipped else file_result.get("error", "识别失败"),
            ]
        )

    success_fill = PatternFill("solid", fgColor="E7F4ED")
    failed_fill = PatternFill("solid", fgColor="FCE8E8")
    skipped_fill = PatternFill("solid", fgColor="F2F4F7")
    for row_cells in ws_status.iter_rows(min_row=2, max_row=ws_status.max_row, max_col=len(status_headers)):
        status_value = row_cells[2].value
        for cell in row_cells:
            cell.font = body_font
            cell.alignment = left_alignment if cell.column == 6 else center_alignment
            cell.border = border
            if status_value == "成功":
                cell.fill = success_fill
            elif status_value == "未选择":
                cell.fill = skipped_fill
            else:
                cell.fill = failed_fill

    ws_status.freeze_panes = "A2"
    status_widths = {"A": 10, "B": 18, "C": 12, "D": 12, "E": 14, "F": 60}
    for column_letter, width in status_widths.items():
        ws_status.column_dimensions[column_letter].width = width

    for worksheet in wb.worksheets:
        worksheet.row_dimensions[1].height = 26
        for row_idx in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row_idx].height = 22

    wb.save(excel_path)


def _style_header_row(
    worksheet,
    headers,
    header_gray,
    header_blue,
    header_font,
    header_white_font,
    alignment,
    border,
    blue_headers=None,
):
    blue_header_set = blue_headers or {"操作方式", "部门"}
    for col_idx, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        use_blue = header in blue_header_set
        cell.fill = header_blue if use_blue else header_gray
        cell.font = header_white_font if use_blue else header_font
        cell.alignment = alignment
        cell.border = border


def _merge_flowchart_groups(ws):
    if ws.max_row <= 2:
        return

    start_row = 2
    previous_key = (ws.cell(row=2, column=1).value, ws.cell(row=2, column=2).value)
    for row_idx in range(3, ws.max_row + 2):
        current_key = (
            ws.cell(row=row_idx, column=1).value if row_idx <= ws.max_row else None,
            ws.cell(row=row_idx, column=2).value if row_idx <= ws.max_row else None,
        )
        if current_key == previous_key:
            continue

        end_row = row_idx - 1
        if end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
            ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=start_row, column=2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        start_row = row_idx
        previous_key = current_key


def _adjust_sheet_widths(workbook):
    for worksheet in workbook.worksheets:
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
