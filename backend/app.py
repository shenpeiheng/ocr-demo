"""
OCR工业图片识别系统 - 后端API服务
支持上传工业图纸图片和PDF文件，识别内容并输出坐标和内容
支持PaddleOCR和OpenAI VL（ModelScope API）两种OCR引擎
"""

import os
import json
import uuid
import io
import base64
import tempfile
import shutil
import re
import requests
import threading
from datetime import datetime
from flask import Flask, request, jsonify, send_file, send_from_directory, render_template
from flask_cors import CORS
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl import Workbook
from PIL import Image, ImageDraw, ImageFont
import numpy as np
import cv2

# 导入配置和处理器
from config import Config
from ocr_processor import OCRProcessor, create_ocr_processor
from pdf_processor import create_pdf_processor
from prompt_manager import prompt_manager, get_prompt
from markdown_formatter import markdown_formatter, format_results, save_markdown
from image_utils import preprocess_image_for_ocr

# 初始化Flask应用
app = Flask(__name__, static_folder='../frontend', static_url_path='')

# 根据配置决定是否启用CORS
if Config.ENABLE_CORS:
    CORS(app)  # 允许跨域请求

# 应用配置
app.config['MAX_CONTENT_LENGTH'] = Config.MAX_CONTENT_LENGTH
app.config['UPLOAD_FOLDER'] = Config.UPLOAD_FOLDER
app.config['ALLOWED_EXTENSIONS'] = Config.ALLOWED_EXTENSIONS
app.config['SECRET_KEY'] = Config.SECRET_KEY

# 确保上传目录存在
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# 初始化OCR处理器（根据环境变量自动选择引擎）
ocr_processor = OCRProcessor()

# 初始化PDF处理器
pdf_processor = create_pdf_processor()

def allowed_file(filename):
    """检查文件扩展名是否允许"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def is_pdf_file(filename):
    """检查文件是否为PDF格式"""
    ext = os.path.splitext(filename)[1].lower()
    return ext == '.pdf'

@app.route('/')
def index():
    """前端页面"""
    return send_from_directory('../frontend', 'index.html')

@app.route('/api')
def api_index():
    """API根路径"""
    engine_info = ocr_processor.get_engine_info()
    
    return jsonify({
        'name': 'OCR工业图片识别系统',
        'version': '3.0.0',
        'description': '工业图纸和PDF文件OCR识别API服务，支持PaddleOCR和OpenAI VL',
        'supported_formats': list(app.config['ALLOWED_EXTENSIONS']),
        'pdf_support': pdf_processor.initialized,
        'ocr_engine': {
            'current': engine_info['current_engine'],
            'type': engine_info['engine_type'],
            'paddleocr_available': engine_info['paddleocr_available'],
            'openai_vl_available': engine_info['openai_vl_available']
        },
        'endpoints': {
            '/api/upload': '上传文件（支持图片和PDF）',
            '/api/process': '处理已上传的文件（自动识别文件类型）',
            '/api/process/openai_vl': '使用OpenAI VL处理图片',
            '/api/process/paddleocr': '使用PaddleOCR处理图片',
            '/api/prompts': '获取可用提示词列表',
            '/api/results/<filename>': '获取识别结果',
            '/api/download/excel/<filename>': '下载Excel格式结果',
            '/api/download/json/<filename>': '下载JSON格式结果',
            '/api/download/markdown/<filename>': '下载Markdown格式结果',
            '/api/pdf/info/<filename>': '获取PDF文件信息',
            '/api/pdf/extract/<filename>': '直接提取PDF文本'
        }
    })

@app.route('/api/upload', methods=['POST'])
def upload_file():
    """上传图片文件"""
    if 'file' not in request.files:
        return jsonify({'error': '没有文件部分'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': '没有选择文件'}), 400
    
    if file and allowed_file(file.filename):
        # 生成唯一文件名
        # 从原始文件名获取扩展名（在secure_filename之前）
        original_filename = file.filename
        # 使用os.path.splitext获取扩展名，更安全
        file_ext = os.path.splitext(original_filename)[1].lower()
        if file_ext:
            file_ext = file_ext[1:]  # 去掉点号，例如 ".pdf" -> "pdf"
        else:
            file_ext = "pdf"  # 默认扩展名
        
        # 生成唯一文件名
        unique_filename = f"{uuid.uuid4().hex}.{file_ext}"
        
        # 保存文件
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
        file.save(filepath)
        
        # 返回文件信息
        return jsonify({
            'success': True,
            'message': '文件上传成功',
            'filename': unique_filename,
            'original_filename': original_filename,
            'filepath': filepath,
            'upload_time': datetime.now().isoformat()
        })
    
    return jsonify({'error': '文件类型不支持'}), 400

@app.route('/api/process', methods=['POST'])
def process_file():
    """处理文件（图片或PDF）并进行OCR识别"""
    data = request.json
    if not data or 'filename' not in data:
        return jsonify({'error': '缺少文件名参数'}), 400
    
    filename = data['filename']
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    
    if not os.path.exists(filepath):
        return jsonify({'error': '文件不存在'}), 404
    
    try:
        # 检查文件类型
        if is_pdf_file(filename):
            # PDF文件处理
            return process_pdf_file(filename, filepath, data)
        else:
            # 图片文件处理
            return process_image_file(filename, filepath, data)
        
    except Exception as e:
        return jsonify({'error': f'处理失败: {str(e)}'}), 500

@app.route('/api/process/openai_vl', methods=['POST'])
def process_with_openai_vl():
    """使用OpenAI VL处理图片"""
    data = request.json
    if not data or 'filename' not in data:
        return jsonify({'error': '缺少文件名参数'}), 400
    
    filename = data['filename']
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    
    if not os.path.exists(filepath):
        return jsonify({'error': '文件不存在'}), 404
    
    if is_pdf_file(filename):
        return jsonify({'error': 'OpenAI VL暂不支持PDF文件，请先转换为图片'}), 400
    
    try:
        # 获取提示词
        prompt_name = data.get('prompt', 'mechanical_drawing_standard')
        custom_prompt = data.get('custom_prompt')
        
        if custom_prompt:
            prompt = custom_prompt
        else:
            prompt = get_prompt(prompt_name)
        
        # 图片预处理：检查尺寸并调整，确保不超过2048x2048（ModelScope API限制）
        preprocessed_path = preprocess_image_for_ocr(filepath, target_size=990, max_size=2048)
        use_preprocessed = preprocessed_path != filepath
        
        # 使用OpenAI VL处理图片
        results = ocr_processor.process_image_with_engine(preprocessed_path, 'openai_vl', prompt)
        
        # 在结果中添加预处理信息
        if use_preprocessed:
            results['preprocessed_image'] = os.path.basename(preprocessed_path)
            results['original_image'] = filename
            results['image_preprocessed'] = True
        else:
            results['image_preprocessed'] = False
        
        # 保存结果到JSON文件
        result_filename = f"result_openai_vl_{os.path.splitext(filename)[0]}.json"
        result_path = os.path.join(app.config['UPLOAD_FOLDER'], result_filename)
        
        with open(result_path, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        
        # 生成Markdown文件
        markdown_filename = f"result_openai_vl_{os.path.splitext(filename)[0]}.md"
        markdown_path = os.path.join(app.config['UPLOAD_FOLDER'], markdown_filename)
        markdown_content = markdown_formatter.format_ocr_results(results, include_raw=data.get('include_raw', False))
        save_markdown(markdown_content, markdown_path)
        
        # 生成Excel文件
        excel_filename = f"result_openai_vl_{os.path.splitext(filename)[0]}.xlsx"
        excel_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_filename)
        generate_excel(results, excel_path)
        
        return jsonify({
            'success': True,
            'message': 'OpenAI VL处理成功',
            'filename': filename,
            'file_type': 'image',
            'engine': 'openai_vl',
            'prompt_used': prompt_name,
            'results': results,
            'result_files': {
                'json': result_filename,
                'excel': excel_filename,
                'markdown': markdown_filename
            }
        })
        
    except Exception as e:
        return jsonify({'error': f'OpenAI VL处理失败: {str(e)}'}), 500

@app.route('/api/process/paddleocr', methods=['POST'])
def process_with_paddleocr():
    """使用PaddleOCR处理图片"""
    data = request.json
    if not data or 'filename' not in data:
        return jsonify({'error': '缺少文件名参数'}), 400
    
    filename = data['filename']
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    
    if not os.path.exists(filepath):
        return jsonify({'error': '文件不存在'}), 404
    
    try:
        # 图片预处理：检查尺寸并调整到990x990（如果需要）
        preprocessed_path = preprocess_image_for_ocr(filepath, target_size=990)
        use_preprocessed = preprocessed_path != filepath
        
        # 使用PaddleOCR处理图片
        results = ocr_processor.process_image_with_engine(preprocessed_path, 'paddleocr')
        
        # 在结果中添加预处理信息
        if use_preprocessed:
            results['preprocessed_image'] = os.path.basename(preprocessed_path)
            results['original_image'] = filename
            results['image_preprocessed'] = True
        else:
            results['image_preprocessed'] = False
        
        # 保存结果到JSON文件
        result_filename = f"result_paddleocr_{os.path.splitext(filename)[0]}.json"
        result_path = os.path.join(app.config['UPLOAD_FOLDER'], result_filename)
        
        with open(result_path, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        
        # 生成Excel文件
        excel_filename = f"result_paddleocr_{os.path.splitext(filename)[0]}.xlsx"
        excel_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_filename)
        generate_excel(results, excel_path)
        
        return jsonify({
            'success': True,
            'message': 'PaddleOCR处理成功',
            'filename': filename,
            'file_type': 'image',
            'engine': 'paddleocr',
            'results': results,
            'result_files': {
                'json': result_filename,
                'excel': excel_filename
            }
        })
        
    except Exception as e:
        return jsonify({'error': f'PaddleOCR处理失败: {str(e)}'}), 500

@app.route('/api/process/custom', methods=['POST'])
def process_with_custom_prompt():
    """使用自定义提示词处理图片，返回原始文本结果及带标注的结果图片"""
    data = request.json
    if not data or 'filename' not in data:
        return jsonify({'error': '缺少文件名参数'}), 400
    
    if 'prompt' not in data or not data['prompt'].strip():
        return jsonify({'error': '缺少提示词参数'}), 400
    
    filename = data['filename']
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    
    if not os.path.exists(filepath):
        return jsonify({'error': '文件不存在'}), 404
    
    if is_pdf_file(filename):
        return jsonify({'error': '自定义提示词暂不支持PDF文件'}), 400
    
    try:
        prompt = data['prompt']
        
        # 追加坐标格式要求：要求模型在输出末尾附加Markdown表格
        # 使用 (x1, y1, x2, y2) 格式，避免HTML标签冲突
        coord_note = """


        【重要输出要求】
        请在回答的最后，严格按照以下Markdown表格格式列出所有识别到的信息项及其位置坐标：

        | 序号 | 内容 | 类型 | 区域 | 坐标 |
        |------|------|------|------|------|
        | 1 | 识别出的文本 | 类型名称 | 区域描述 | (x1, y1, x2, y2) |

        坐标请使用 (左上角x, 左上角y, 右下角x, 右下角y) 格式，坐标范围为 [0, 1000]，所有数值为整数。"""
        prompt = prompt + coord_note
        
        # 图片预处理：检查尺寸并调整，确保不超过2048x2048（ModelScope API限制）
        preprocessed_path = preprocess_image_for_ocr(filepath, target_size=990, max_size=2048)
        
        # 使用OpenAI VL处理图片（强制使用openai_vl引擎）
        results = ocr_processor.process_image_with_engine(preprocessed_path, 'openai_vl', prompt)
        
        # 提取原始响应文本
        raw_text = results.get('raw_response', '')
        # 如果raw_response被截断，尝试从results中获取完整内容
        if not raw_text and 'text_items' in results:
            # 从text_items构建文本
            text_parts = []
            for item in results.get('text_items', []):
                text_parts.append(f"{item.get('text', '')}")
            raw_text = '\n'.join(text_parts)
        
        # 生成带标注框的结果图片
        # 注意：坐标是基于预处理后的图片尺寸解析的，所以要用预处理后的图片来画框
        result_image_base64 = None
        text_items = results.get('text_items', [])
        draw_source = preprocessed_path if os.path.exists(preprocessed_path) else filepath
        if os.path.exists(draw_source):
            try:
                result_image_base64 = draw_ocr_boxes(draw_source, text_items)
            except Exception as draw_err:
                print(f"[WARN] 绘制标注框失败: {draw_err}")
        
        # 如果绘制失败或text_items为空，至少返回原始图片
        if result_image_base64 is None and os.path.exists(filepath):
            try:
                with open(filepath, 'rb') as f:
                    img_data = f.read()
                result_image_base64 = f'data:image/jpeg;base64,{base64.b64encode(img_data).decode("utf-8")}'
            except Exception as e:
                print(f"[WARN] 读取原始图片失败: {e}")
        
        # 构建JSON结果（格式化输出）
        json_result = format_results_as_json(results)
        
        return jsonify({
            'success': True,
            'message': '自定义提示词处理成功',
            'filename': filename,
            'engine': 'openai_vl',
            'raw_text': raw_text,
            'results': results,
            'result_image': result_image_base64,
            'json_result': json_result
        })
        
    except Exception as e:
        return jsonify({'error': f'自定义提示词处理失败: {str(e)}'}), 500

@app.route('/api/prompts', methods=['GET'])
def get_prompts():
    """获取可用提示词列表"""
    try:
        prompts = prompt_manager.get_all_prompts()
        
        # 只返回名称和预览
        prompt_list = []
        for name, content in prompts.items():
            prompt_list.append({
                'name': name,
                'preview': content[:100] + "..." if len(content) > 100 else content,
                'length': len(content)
            })
        
        return jsonify({
            'success': True,
            'prompts': prompt_list,
            'total': len(prompt_list)
        })
    except Exception as e:
        return jsonify({'error': f'获取提示词失败: {str(e)}'}), 500

@app.route('/api/prompts/<prompt_name>', methods=['GET'])
def get_prompt_detail(prompt_name):
    """获取特定提示词的详细内容"""
    try:
        prompt = prompt_manager.get_prompt(prompt_name)
        
        return jsonify({
            'success': True,
            'name': prompt_name,
            'content': prompt,
            'length': len(prompt)
        })
    except Exception as e:
        return jsonify({'error': f'获取提示词失败: {str(e)}'}), 500

def process_image_file(filename, filepath, data):
    """处理图片文件"""
    try:
        # 获取处理参数
        engine = data.get('engine', 'auto')  # auto, paddleocr, openai_vl
        prompt_name = data.get('prompt', 'mechanical_drawing_standard')
        custom_prompt = data.get('custom_prompt')
        
        # 图片预处理：检查尺寸并调整，确保不超过2048x2048（ModelScope API限制）
        preprocessed_path = preprocess_image_for_ocr(filepath, target_size=990, max_size=2048)
        use_preprocessed = preprocessed_path != filepath
        
        # 选择处理方式
        if engine == 'openai_vl':
            # 使用OpenAI VL
            if custom_prompt:
                prompt = custom_prompt
            else:
                prompt = get_prompt(prompt_name)
            
            results = ocr_processor.process_image_with_engine(preprocessed_path, 'openai_vl', prompt)
            result_prefix = 'result_openai_vl'
        elif engine == 'paddleocr':
            # 使用PaddleOCR
            results = ocr_processor.process_image_with_engine(preprocessed_path, 'paddleocr')
            result_prefix = 'result_paddleocr'
        else:
            # 自动选择（使用初始化时的引擎）
            if custom_prompt:
                results = ocr_processor.process_image(preprocessed_path, custom_prompt)
            else:
                results = ocr_processor.process_image(preprocessed_path)
            result_prefix = 'result'
        
        # 在结果中添加预处理信息
        if use_preprocessed:
            results['preprocessed_image'] = os.path.basename(preprocessed_path)
            results['original_image'] = filename
            results['image_preprocessed'] = True
        else:
            results['image_preprocessed'] = False
        
        # 保存结果到JSON文件
        result_filename = f"{result_prefix}_{os.path.splitext(filename)[0]}.json"
        result_path = os.path.join(app.config['UPLOAD_FOLDER'], result_filename)
        
        with open(result_path, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        
        # 生成Excel文件
        excel_filename = f"{result_prefix}_{os.path.splitext(filename)[0]}.xlsx"
        excel_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_filename)
        generate_excel(results, excel_path)
        
        # 如果是OpenAI VL，额外生成Markdown文件
        markdown_filename = None
        if engine == 'openai_vl' or (engine == 'auto' and ocr_processor.current_engine == 'openai_vl'):
            markdown_filename = f"{result_prefix}_{os.path.splitext(filename)[0]}.md"
            markdown_path = os.path.join(app.config['UPLOAD_FOLDER'], markdown_filename)
            markdown_content = markdown_formatter.format_ocr_results(results, include_raw=data.get('include_raw', False))
            save_markdown(markdown_content, markdown_path)
        
        # 构建响应
        response = {
            'success': True,
            'message': '图片处理成功',
            'filename': filename,
            'file_type': 'image',
            'engine': results.get('ocr_engine', 'unknown'),
            'results': results,
            'result_files': {
                'json': result_filename,
                'excel': excel_filename
            }
        }
        
        if markdown_filename:
            response['result_files']['markdown'] = markdown_filename
        
        return jsonify(response)
        
    except Exception as e:
        return jsonify({'error': f'图片处理失败: {str(e)}'}), 500

def process_pdf_file(filename, filepath, data):
    """处理PDF文件"""
    try:
        # 检查PDF处理器是否可用
        if not pdf_processor.initialized:
            return jsonify({
                'success': False,
                'error': 'PDF处理器未初始化，请安装PDF处理依赖',
                'suggestion': '请运行: pip install pdf2image PyMuPDF pdfplumber'
            }), 500
        
        # 获取处理参数
        max_pages = data.get('max_pages', Config.PDF_MAX_PAGES)
        dpi = data.get('dpi', Config.PDF_DPI)
        engine = data.get('engine', 'auto')
        
        # 创建PDF图像输出目录（在上传目录中）
        pdf_base_name = os.path.splitext(filename)[0]
        images_dir = os.path.join(app.config['UPLOAD_FOLDER'], f"images_{pdf_base_name}")
        os.makedirs(images_dir, exist_ok=True)
        
        # 处理PDF文件，指定输出目录
        pdf_result = pdf_processor.process_pdf_with_ocr(
            filepath,
            ocr_processor,
            dpi=dpi,
            max_pages=max_pages,
            output_dir=images_dir
        )
        
        if not pdf_result['success']:
            return jsonify({
                'success': False,
                'error': pdf_result.get('error', 'PDF处理失败'),
                'file_type': 'pdf'
            }), 500
        
        # 保存结果到JSON文件
        result_filename = f"result_{pdf_base_name}.json"
        result_path = os.path.join(app.config['UPLOAD_FOLDER'], result_filename)
        
        with open(result_path, 'w', encoding='utf-8') as f:
            json.dump(pdf_result, f, ensure_ascii=False, indent=2)
        
        # 生成Excel文件（包含所有页面的结果）
        excel_filename = f"result_{pdf_base_name}.xlsx"
        excel_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_filename)
        generate_pdf_excel(pdf_result, excel_path)
        
        # 不再清理图像文件，因为现在保存在上传目录中
        # 但需要更新图像路径，使其可以通过API访问
        if 'conversion_info' in pdf_result and 'image_paths' in pdf_result['conversion_info']:
            # 将绝对路径转换为相对路径
            image_paths = pdf_result['conversion_info']['image_paths']
            relative_image_paths = []
            for img_path in image_paths:
                if os.path.exists(img_path):
                    # 计算相对于上传目录的路径
                    rel_path = os.path.relpath(img_path, app.config['UPLOAD_FOLDER'])
                    relative_image_paths.append(rel_path)
            
            # 保存相对路径到结果中
            pdf_result['conversion_info']['relative_image_paths'] = relative_image_paths
        
        return jsonify({
            'success': True,
            'message': 'PDF处理成功',
            'filename': filename,
            'file_type': 'pdf',
            'results': pdf_result,
            'result_files': {
                'json': result_filename,
                'excel': excel_filename
            }
        })
        
    except Exception as e:
        return jsonify({'error': f'PDF处理失败: {str(e)}'}), 500

@app.route('/api/results/<filename>')
def get_results(filename):
    """获取识别结果"""
    # 尝试多种可能的结果文件名
    possible_filenames = [
        f"result_{os.path.splitext(filename)[0]}.json",
        f"result_openai_vl_{os.path.splitext(filename)[0]}.json",
        f"result_paddleocr_{os.path.splitext(filename)[0]}.json"
    ]
    
    result_path = None
    for possible_filename in possible_filenames:
        test_path = os.path.join(app.config['UPLOAD_FOLDER'], possible_filename)
        if os.path.exists(test_path):
            result_path = test_path
            break
    
    if not result_path:
        return jsonify({'error': '结果文件不存在'}), 404
    
    try:
        with open(result_path, 'r', encoding='utf-8') as f:
            results = json.load(f)
        
        return jsonify({
            'success': True,
            'filename': filename,
            'result_file': os.path.basename(result_path),
            'results': results
        })
    except Exception as e:
        return jsonify({'error': f'读取结果失败: {str(e)}'}), 500

@app.route('/api/download/excel/<filename>')
def download_excel(filename):
    """下载Excel格式结果"""
    # 尝试多种可能的Excel文件名
    possible_filenames = [
        f"result_{os.path.splitext(filename)[0]}.xlsx",
        f"result_openai_vl_{os.path.splitext(filename)[0]}.xlsx",
        f"result_paddleocr_{os.path.splitext(filename)[0]}.xlsx"
    ]
    
    excel_path = None
    for possible_filename in possible_filenames:
        test_path = os.path.join(app.config['UPLOAD_FOLDER'], possible_filename)
        if os.path.exists(test_path):
            excel_path = test_path
            break
    
    if not excel_path:
        return jsonify({'error': 'Excel文件不存在'}), 404
    
    return send_file(excel_path, as_attachment=True, download_name=f"ocr_result_{filename}.xlsx")

@app.route('/api/download/json/<filename>')
def download_json(filename):
    """下载JSON格式结果"""
    # 尝试多种可能的JSON文件名
    possible_filenames = [
        f"result_{os.path.splitext(filename)[0]}.json",
        f"result_openai_vl_{os.path.splitext(filename)[0]}.json",
        f"result_paddleocr_{os.path.splitext(filename)[0]}.json"
    ]
    
    json_path = None
    for possible_filename in possible_filenames:
        test_path = os.path.join(app.config['UPLOAD_FOLDER'], possible_filename)
        if os.path.exists(test_path):
            json_path = test_path
            break
    
    if not json_path:
        return jsonify({'error': 'JSON文件不存在'}), 404
    
    return send_file(json_path, as_attachment=True, download_name=f"ocr_result_{filename}.json")

@app.route('/api/download/markdown/<filename>')
def download_markdown(filename):
    """下载Markdown格式结果"""
    # 尝试多种可能的Markdown文件名
    possible_filenames = [
        f"result_{os.path.splitext(filename)[0]}.md",
        f"result_openai_vl_{os.path.splitext(filename)[0]}.md"
    ]
    
    markdown_path = None
    for possible_filename in possible_filenames:
        test_path = os.path.join(app.config['UPLOAD_FOLDER'], possible_filename)
        if os.path.exists(test_path):
            markdown_path = test_path
            break
    
    if not markdown_path:
        return jsonify({'error': 'Markdown文件不存在'}), 404
    
    return send_file(markdown_path, as_attachment=True, download_name=f"ocr_result_{filename}.md")

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    """访问上传的文件"""
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

# 全局缓存 RetinaFace 模型（线程安全）
_face_detection_lock = threading.Lock()
_face_detection_pipeline = None

def _get_face_detection_pipeline():
    """获取或初始化 ModelScope RetinaFace 人脸检测 pipeline（延迟加载，线程安全）"""
    global _face_detection_pipeline
    if _face_detection_pipeline is None:
        with _face_detection_lock:
            if _face_detection_pipeline is None:
                try:
                    from modelscope.pipelines import pipeline
                    from modelscope.utils.constant import Tasks
                    _face_detection_pipeline = pipeline(
                        Tasks.face_detection,
                        model='iic/cv_resnet50_face-detection_retinaface'
                    )
                except Exception as e:
                    print(f"[FaceDetection] 加载 RetinaFace 模型失败: {e}")
                    _face_detection_pipeline = None
    return _face_detection_pipeline

@app.route('/api/face_detection', methods=['POST'])
def face_detection():
    """
    人脸检测 API
    使用 ModelScope RetinaFace 模型 (iic/cv_resnet50_face-detection_retinaface)
    高精度人脸检测，支持密集人脸场景
    """
    if 'file' not in request.files:
        return jsonify({'error': '没有文件部分'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': '没有选择文件'}), 400
    
    try:
        # 保存上传的图片到临时文件（RetinaFace pipeline 需要文件路径）
        image_bytes = file.read()
        nparr = np.frombuffer(image_bytes, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        
        if img is None:
            return jsonify({'error': '无法解码图片'}), 400
        
        img_height, img_width = img.shape[:2]
        
        # 保存临时文件供 RetinaFace 使用
        temp_dir = tempfile.gettempdir()
        temp_path = os.path.join(temp_dir, f'face_detection_{uuid.uuid4().hex}.jpg')
        cv2.imwrite(temp_path, img)
        
        try:
            # 使用 RetinaFace 模型进行人脸检测
            pipeline = _get_face_detection_pipeline()
            
            if pipeline is not None:
                result = pipeline(temp_path)
                
                scores = result.get('scores', [])
                boxes = result.get('boxes', [])
                
                faces = []
                for i, box in enumerate(boxes):
                    if len(box) >= 4:
                        x1, y1, x2, y2 = box[:4]
                        confidence = float(scores[i]) if i < len(scores) else 0.0
                        
                        # 确保坐标在图片范围内
                        x1 = max(0, int(x1))
                        y1 = max(0, int(y1))
                        x2 = min(img_width, int(x2))
                        y2 = min(img_height, int(y2))
                        
                        faces.append({
                            'bbox': {
                                'x1': float(x1),
                                'y1': float(y1),
                                'x2': float(x2),
                                'y2': float(y2),
                                'width': float(x2 - x1),
                                'height': float(y2 - y1)
                            },
                            'confidence': round(confidence, 4)
                        })
                
                detector_name = 'ModelScope RetinaFace'
            else:
                # Fallback: 使用 OpenCV DNN (ResNet SSD)
                model_dir = os.path.join(os.path.dirname(__file__), 'models')
                prototxt = os.path.join(model_dir, 'deploy.prototxt')
                caffemodel = os.path.join(model_dir, 'res10_300x300_ssd_iter_140000.caffemodel')
                
                if os.path.exists(prototxt) and os.path.exists(caffemodel):
                    net = cv2.dnn.readNetFromCaffe(prototxt, caffemodel)
                    blob = cv2.dnn.blobFromImage(
                        cv2.resize(img, (300, 300)),
                        1.0, (300, 300), (104.0, 177.0, 123.0)
                    )
                    net.setInput(blob)
                    detections = net.forward()
                    
                    faces = []
                    for i in range(detections.shape[2]):
                        confidence = float(detections[0, 0, i, 2])
                        if confidence > 0.5:
                            box = detections[0, 0, i, 3:7] * np.array(
                                [img_width, img_height, img_width, img_height]
                            )
                            x1, y1, x2, y2 = box.astype(int)
                            x1 = max(0, x1); y1 = max(0, y1)
                            x2 = min(img_width, x2); y2 = min(img_height, y2)
                            faces.append({
                                'bbox': {'x1': float(x1), 'y1': float(y1),
                                        'x2': float(x2), 'y2': float(y2),
                                        'width': float(x2 - x1), 'height': float(y2 - y1)},
                                'confidence': round(confidence, 4)
                            })
                    detector_name = 'OpenCV DNN (ResNet SSD)'
                else:
                    # Fallback: Haar Cascade
                    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                    cascade_path = cv2.data.haarcascades + 'haarcascade_frontalface_default.xml'
                    face_cascade = cv2.CascadeClassifier(cascade_path)
                    faces_data = face_cascade.detectMultiScale(gray, 1.1, 5, minSize=(30, 30))
                    faces = []
                    for (x, y, w, h) in faces_data:
                        faces.append({
                            'bbox': {'x1': float(x), 'y1': float(y),
                                    'x2': float(x + w), 'y2': float(y + h),
                                    'width': float(w), 'height': float(h)},
                            'confidence': 0.95
                        })
                    detector_name = 'OpenCV Haar Cascade'
        finally:
            # 清理临时文件
            try:
                os.remove(temp_path)
            except:
                pass
        
        # 在图片上绘制检测框
        img_draw = img.copy()
        for face in faces:
            bbox = face['bbox']
            x1 = int(bbox['x1'])
            y1 = int(bbox['y1'])
            x2 = int(bbox['x2'])
            y2 = int(bbox['y2'])
            
            # 绘制矩形框（绿色）
            cv2.rectangle(img_draw, (x1, y1), (x2, y2), (0, 255, 0), 2)
            
            # 绘制置信度标签
            label = f"Face {face['confidence']:.2f}"
            (label_w, label_h), _ = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.6, 2)
            cv2.rectangle(img_draw, (x1, y1 - label_h - 10), (x1 + label_w + 10, y1), (0, 255, 0), -1)
            cv2.putText(img_draw, label, (x1 + 5, y1 - 5),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)
        
        # 将标注后的图片编码为 base64
        _, buffer_draw = cv2.imencode('.jpg', img_draw, [cv2.IMWRITE_JPEG_QUALITY, 90])
        img_result_base64 = base64.b64encode(buffer_draw).decode('utf-8')
        
        return jsonify({
            'success': True,
            'face_count': len(faces),
            'faces': faces,
            'image_width': img_width,
            'image_height': img_height,
            'result_image': f'data:image/jpeg;base64,{img_result_base64}',
            'detector': detector_name,
            'model_ref': 'https://modelscope.cn/models/iic/cv_resnet50_face-detection_retinaface'
        })
        
    except Exception as e:
        return jsonify({'error': f'人脸检测失败: {str(e)}'}), 500

@app.route('/api/pdf/info/<filename>')
def get_pdf_info(filename):
    """获取PDF文件信息"""
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    
    if not os.path.exists(filepath):
        return jsonify({'error': '文件不存在'}), 404
    
    if not is_pdf_file(filename):
        return jsonify({'error': '文件不是PDF格式'}), 400
    
    if not pdf_processor.initialized:
        return jsonify({
            'success': False,
            'error': 'PDF处理器未初始化',
            'suggestion': '请安装PDF处理依赖: pip install pdf2image PyMuPDF pdfplumber'
        }), 500
    
    try:
        pdf_info = pdf_processor._get_pdf_info(filepath)
        return jsonify({
            'success': True,
            'filename': filename,
            'pdf_info': pdf_info
        })
    except Exception as e:
        return jsonify({'error': f'获取PDF信息失败: {str(e)}'}), 500

@app.route('/api/pdf/extract/<filename>')
def extract_pdf_text(filename):
    """直接提取PDF文本（不进行OCR）"""
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    
    if not os.path.exists(filepath):
        return jsonify({'error': '文件不存在'}), 404
    
    if not is_pdf_file(filename):
        return jsonify({'error': '文件不是PDF格式'}), 400
    
    if not pdf_processor.initialized:
        return jsonify({
            'success': False,
            'error': 'PDF处理器未初始化',
            'suggestion': '请安装PDF处理依赖: pip install pdf2image PyMuPDF pdfplumber'
        }), 500
    
    try:
        # 获取查询参数
        first_page = request.args.get('first_page', default=None, type=int)
        last_page = request.args.get('last_page', default=None, type=int)
        
        # 提取文本
        text_result = pdf_processor.extract_text_from_pdf(
            filepath,
            first_page=first_page,
            last_page=last_page
        )
        
        if text_result['success']:
            return jsonify({
                'success': True,
                'filename': filename,
                'result': text_result
            })
        else:
            return jsonify({
                'success': False,
                'error': text_result.get('error', '文本提取失败')
            }), 500
            
    except Exception as e:
        return jsonify({'error': f'提取PDF文本失败: {str(e)}'}), 500

@app.route('/api/pdf/images/<filename>/<int:page>')
def get_pdf_image(filename, page):
    """获取PDF页面图像"""
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    
    if not os.path.exists(filepath):
        return jsonify({'error': '文件不存在'}), 404
    
    if not is_pdf_file(filename):
        return jsonify({'error': '文件不是PDF格式'}), 400
    
    # 查找图像文件
    pdf_base_name = os.path.splitext(filename)[0]
    images_dir = os.path.join(app.config['UPLOAD_FOLDER'], f"images_{pdf_base_name}")
    
    # 可能的图像文件名格式
    image_patterns = [
        f"page_{page:03d}.png",
        f"page_{page}.png",
        f"{page}.png",
        f"page_{page:03d}.jpg",
        f"page_{page}.jpg",
        f"{page}.jpg"
    ]
    
    image_path = None
    for pattern in image_patterns:
        test_path = os.path.join(images_dir, pattern)
        if os.path.exists(test_path):
            image_path = test_path
            break
    
    if not image_path:
        # 如果找不到图像文件，尝试从PDF重新生成
        try:
            if not pdf_processor.initialized:
                return jsonify({'error': 'PDF处理器未初始化'}), 500
            
            # 生成单页图像
            conversion_result = pdf_processor.convert_pdf_to_images(
                filepath,
                dpi=200,
                first_page=page,
                last_page=page,
                output_dir=images_dir
            )
            
            if conversion_result['success'] and conversion_result['image_paths']:
                image_path = conversion_result['image_paths'][0]
            else:
                return jsonify({'error': '无法生成PDF页面图像'}), 500
        except Exception as e:
            return jsonify({'error': f'生成图像失败: {str(e)}'}), 500
    
    # 返回图像文件
    return send_file(image_path, mimetype='image/png')

@app.route('/api/pdf/images/list/<filename>')
def list_pdf_images(filename):
    """列出PDF所有页面图像"""
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    
    if not os.path.exists(filepath):
        return jsonify({'error': '文件不存在'}), 404
    
    if not is_pdf_file(filename):
        return jsonify({'error': '文件不是PDF格式'}), 400
    
    # 查找图像目录
    pdf_base_name = os.path.splitext(filename)[0]
    images_dir = os.path.join(app.config['UPLOAD_FOLDER'], f"images_{pdf_base_name}")
    
    if not os.path.exists(images_dir):
        return jsonify({
            'success': True,
            'filename': filename,
            'images': [],
            'message': '图像目录不存在，PDF可能尚未处理'
        })
    
    # 列出所有PNG和JPG文件
    image_files = []
    for ext in ['.png', '.jpg', '.jpeg']:
        for file in os.listdir(images_dir):
            if file.lower().endswith(ext):
                # 提取页码
                page_match = None
                import re
                # 匹配 page_001.png 或 page_1.png 或 1.png 等格式
                match = re.search(r'page_(\d+)', file)
                if match:
                    page_num = int(match.group(1))
                else:
                    match = re.search(r'(\d+)\.', file)
                    if match:
                        page_num = int(match.group(1))
                    else:
                        page_num = 0
                
                image_files.append({
                    'filename': file,
                    'page': page_num,
                    'url': f"/api/pdf/images/{filename}/{page_num}"
                })
    
    # 按页码排序
    image_files.sort(key=lambda x: x['page'])
    
    return jsonify({
        'success': True,
        'filename': filename,
        'images': image_files,
        'total': len(image_files)
    })

def draw_ocr_boxes(image_path, text_items):
    """
    在图片上绘制OCR识别框和文本标注（支持中文）
    使用统一颜色方案：框线用亮色，标签用白底黑字，清晰易读
    
    Args:
        image_path: 原始图片路径
        text_items: 文本项列表，每项包含 location (left, top, width, height) 和 text
        
    Returns:
        base64编码的结果图片 (data:image/jpeg;base64,...)
    """
    try:
        # 用 PIL 读取图片
        pil_img = Image.open(image_path).convert('RGB')
        img_width, img_height = pil_img.size
        
        # 转换为 OpenCV BGR 格式
        img_cv = cv2.cvtColor(np.array(pil_img), cv2.COLOR_RGB2BGR)
        
        # 尝试加载中文字体（用大号字体让文字更清晰）
        _script_dir = os.path.dirname(os.path.abspath(__file__))
        _project_root = os.path.dirname(_script_dir)
        font_paths = [
            os.path.join(_project_root, 'frontend', 'static', 'fonts', 'NotoSansSC.ttf'),
            'C:/Windows/Fonts/msyh.ttc',
            'C:/Windows/Fonts/simhei.ttf',
            'C:/Windows/Fonts/simsun.ttc',
            'C:/Windows/Fonts/yahei.ttf',
            'C:/Windows/Fonts/msyhbd.ttc',
        ]
        font_large = None  # 用于标签的大字体
        font_small = None  # 用于序号的小字体
        for fp in font_paths:
            if os.path.exists(fp):
                try:
                    font_large = ImageFont.truetype(fp, 20)
                    font_small = ImageFont.truetype(fp, 16)
                    break
                except:
                    continue
        if font_large is None:
            font_large = ImageFont.load_default()
            font_small = ImageFont.load_default()
        
        # 使用统一的颜色方案：红色
        # 框线颜色 (BGR)
        box_color_bgr = (0, 0, 255)      # 红色 (OpenCV BGR)
        box_color_rgb = (255, 0, 0)      # 红色 (PIL RGB)
        
        # 标签背景色：白色半透明效果（用纯白）
        label_bg_bgr = (255, 255, 255)   # 白色 (OpenCV BGR)
        label_text_color = (0, 0, 0)     # 黑色文字
        
        print(f"[draw_ocr_boxes] 图片尺寸: {img_width}x{img_height}, text_items数量: {len(text_items)}")
        
        for i, item in enumerate(text_items):
            location = item.get('location', {})
            left = int(location.get('left', 0))
            top = int(location.get('top', 0))
            w = int(location.get('width', 50))
            h = int(location.get('height', 20))
            
            # 确保坐标在图片范围内
            left = max(0, min(left, img_width - 1))
            top = max(0, min(top, img_height - 1))
            right = min(left + w, img_width)
            bottom = min(top + h, img_height)
            
            if right <= left or bottom <= top:
                continue
            
            text = item.get('text', '')
            
            # ---- 1. 绘制矩形框（亮蓝色，线宽2） ----
            cv2.rectangle(img_cv, (left, top), (right, bottom), box_color_bgr, 2)
            
            # ---- 2. 绘制序号标签（白底黑字，放在框左上角） ----
            label = f"#{i+1}"
            
            # 用 PIL 计算标签尺寸
            pil_tmp = Image.fromarray(cv2.cvtColor(img_cv, cv2.COLOR_BGR2RGB))
            draw_tmp = ImageDraw.Draw(pil_tmp)
            bbox = draw_tmp.textbbox((0, 0), label, font=font_small)
            label_w = bbox[2] - bbox[0] + 12  # 加内边距
            label_h = bbox[3] - bbox[1] + 6   # 加内边距
            
            # 标签位置：框的左上角
            label_x = left
            label_y = top - label_h - 2
            if label_y < 0:
                label_y = top + 2
            
            # 绘制白色标签背景
            cv2.rectangle(img_cv,
                         (label_x, label_y),
                         (label_x + label_w, label_y + label_h),
                         label_bg_bgr, -1)
            # 绘制标签边框（用框线颜色）
            cv2.rectangle(img_cv,
                         (label_x, label_y),
                         (label_x + label_w, label_y + label_h),
                         box_color_bgr, 1)
            
            # 用 PIL 绘制黑色序号文字
            pil_tmp2 = Image.fromarray(cv2.cvtColor(img_cv, cv2.COLOR_BGR2RGB))
            draw_tmp2 = ImageDraw.Draw(pil_tmp2)
            draw_tmp2.text((label_x + 6, label_y + 3), label, fill=label_text_color, font=font_small)
            img_cv = cv2.cvtColor(np.array(pil_tmp2), cv2.COLOR_RGB2BGR)
        
        # 编码输出
        final_img = Image.fromarray(cv2.cvtColor(img_cv, cv2.COLOR_BGR2RGB))
        output_buffer = io.BytesIO()
        final_img.save(output_buffer, format='JPEG', quality=95)
        img_base64 = base64.b64encode(output_buffer.getvalue()).decode('utf-8')
        
        return f'data:image/jpeg;base64,{img_base64}'
        
    except Exception as e:
        print(f"[draw_ocr_boxes] 错误: {e}")
        import traceback
        traceback.print_exc()
        return None


def format_results_as_json(results):
    """
    将OCR结果格式化为结构化的JSON字符串（用于前端展示）
    
    Args:
        results: OCR处理结果字典
        
    Returns:
        格式化的JSON字符串
    """
    try:
        text_items = results.get('text_items', [])
        image_info = results.get('image_info', {})
        analysis = results.get('analysis', {})
        
        formatted = {
            '识别统计': {
                '总识别项数': results.get('total_items', 0),
                '处理时间(秒)': results.get('processing_time', 0),
                'OCR引擎': results.get('ocr_engine', 'unknown'),
                '图片宽度': image_info.get('width', 0),
                '图片高度': image_info.get('height', 0),
            },
            '类型分布': analysis.get('type_distribution', {}),
            '识别详情': []
        }
        
        for i, item in enumerate(text_items):
            location = item.get('location', {})
            detail = {
                '序号': i + 1,
                '内容': item.get('text', ''),
                '类型': item.get('type', 'text'),
                '区域': item.get('region', ''),
                '坐标': {
                    'left': location.get('left', 0),
                    'top': location.get('top', 0),
                    'width': location.get('width', 0),
                    'height': location.get('height', 0)
                },
                '置信度': item.get('confidence', 0)
            }
            formatted['识别详情'].append(detail)
        
        return json.dumps(formatted, ensure_ascii=False, indent=2)
        
    except Exception as e:
        print(f"[format_results_as_json] 错误: {e}")
        return json.dumps({'error': str(e)}, ensure_ascii=False, indent=2)


def generate_excel(results, excel_path):
    """生成Excel文件（用于图片处理结果）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "OCR识别结果"
    
    # 添加表头
    headers = ['序号', '坐标X', '坐标Y', '宽度', '高度', '内容', '置信度', '类型']
    if 'region' in results.get('text_items', [{}])[0]:
        headers.append('区域')
    
    ws.append(headers)
    
    # 添加数据
    for i, item in enumerate(results.get('text_items', []), 1):
        location = item.get('location', {})
        row_data = [
            i,
            location.get('left', 0),
            location.get('top', 0),
            location.get('width', 0),
            location.get('height', 0),
            item.get('text', ''),
            item.get('confidence', 0),
            item.get('type', 'text')
        ]
        
        if 'region' in item:
            row_data.append(item.get('region', ''))
        
        ws.append(row_data)
    
    # 调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(excel_path)

def generate_pdf_excel(pdf_result, excel_path):
    """生成Excel文件（用于PDF处理结果）"""
    wb = Workbook()
    
    # 创建汇总工作表
    ws_summary = wb.active
    ws_summary.title = "PDF处理汇总"
    
    # 添加汇总信息
    ws_summary.append(['PDF文件信息', ''])
    pdf_info = pdf_result.get('pdf_info', {})
    ws_summary.append(['文件名', pdf_info.get('filename', '')])
    ws_summary.append(['总页数', pdf_info.get('total_pages', 0)])
    ws_summary.append(['文件大小', f"{pdf_info.get('file_size', 0)} 字节"])
    
    ws_summary.append([])
    ws_summary.append(['处理信息', ''])
    processing_info = pdf_result.get('processing_info', {})
    ws_summary.append(['DPI', processing_info.get('dpi', 200)])
    ws_summary.append(['最大处理页数', processing_info.get('max_pages', 10)])
    ws_summary.append(['实际处理页数', processing_info.get('actual_pages_processed', 0)])
    
    ws_summary.append([])
    ws_summary.append(['识别结果汇总', ''])
    combined_results = pdf_result.get('combined_results', {})
    ws_summary.append(['总识别项数', combined_results.get('total_items', 0)])
    ws_summary.append(['总处理时间', f"{combined_results.get('total_processing_time', 0)} 秒"])
    
    # 创建页面汇总工作表
    ws_pages = wb.create_sheet(title="页面汇总")
    ws_pages.append(['页码', '识别项数', '处理时间(秒)', '处理状态'])
    
    pages_summary = pdf_result.get('pages_summary', [])
    for page in pages_summary:
        ws_pages.append([
            page.get('page_number', 0),
            page.get('total_items', 0),
            page.get('processing_time', 0),
            '成功' if page.get('success', False) else '失败'
        ])
    
    # 创建详细结果工作表
    ws_details = wb.create_sheet(title="详细识别结果")
    detail_headers = ['页码', '序号', '坐标X', '坐标Y', '宽度', '高度', '内容', '置信度', '类型']
    if 'region' in pdf_result.get('combined_results', {}).get('text_items', [{}])[0]:
        detail_headers.append('区域')
    
    ws_details.append(detail_headers)
    
    text_items = combined_results.get('text_items', [])
    for i, item in enumerate(text_items, 1):
        location = item.get('location', {})
        row_data = [
            item.get('page', 1),
            i,
            location.get('left', 0),
            location.get('top', 0),
            location.get('width', 0),
            location.get('height', 0),
            item.get('text', ''),
            item.get('confidence', 0),
            item.get('type', 'text')
        ]
        
        if 'region' in item:
            row_data.append(item.get('region', ''))
        
        ws_details.append(row_data)
    
    # 调整列宽
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(excel_path)

# ============================================================
# 实时安全帽检测 API
# ============================================================

# 全局缓存安全帽检测模型（线程安全）
_safety_helmet_lock = threading.Lock()
_safety_helmet_pipeline = None

# 最大重试次数
_HELMET_MAX_RETRIES = 3
_HELMET_RETRY_DELAY = 5

def _init_safety_helmet_pipeline():
    """初始化安全帽检测 pipeline（内部方法，已持有锁时调用）"""
    global _safety_helmet_pipeline
    try:
        from modelscope.pipelines import pipeline
        from modelscope.utils.constant import Tasks
        _safety_helmet_pipeline = pipeline(
            Tasks.image_object_detection,
            model='iic/cv_tinynas_object-detection_damoyolo_safety-helmet',
            trust_remote_code=True
        )
        print("[SafetyHelmet] DAMO-YOLO 安全帽检测模型已加载")
        return True
    except Exception as e:
        print(f"[SafetyHelmet] 加载安全帽检测模型失败: {e}")
        _safety_helmet_pipeline = None
        return False

def _get_safety_helmet_pipeline():
    """获取或初始化 ModelScope 安全帽检测 pipeline（延迟加载，线程安全，带重试）"""
    global _safety_helmet_pipeline
    if _safety_helmet_pipeline is None:
        with _safety_helmet_lock:
            if _safety_helmet_pipeline is None:
                for attempt in range(1, _HELMET_MAX_RETRIES + 1):
                    print(f"[SafetyHelmet] 尝试初始化 (第 {attempt}/{_HELMET_MAX_RETRIES} 次)...")
                    if _init_safety_helmet_pipeline():
                        break
                    if attempt < _HELMET_MAX_RETRIES:
                        print(f"[SafetyHelmet] 等待 {_HELMET_RETRY_DELAY} 秒后重试...")
                        import time
                        time.sleep(_HELMET_RETRY_DELAY)
                    else:
                        print(f"[SafetyHelmet] 初始化失败，已重试 {_HELMET_MAX_RETRIES} 次")
    return _safety_helmet_pipeline

def preload_safety_helmet():
    """预加载安全帽检测模型（在应用启动时调用）"""
    print("[SafetyHelmet] 正在预加载安全帽检测模型...")
    pipeline = _get_safety_helmet_pipeline()
    if pipeline is not None:
        print("[SafetyHelmet] 安全帽检测模型预加载成功")
        # 预热模型
        try:
            import numpy as np
            test_img = np.ones((100, 100, 3), dtype=np.uint8) * 255
            pipeline(test_img)
            print("[SafetyHelmet] 安全帽检测模型预热完成")
        except Exception as e:
            print(f"[SafetyHelmet] 模型预热失败（不影响后续使用）: {e}")
    else:
        print("[SafetyHelmet] 安全帽检测模型预加载失败，将在首次请求时重试")

@app.route('/api/safety_helmet_detection', methods=['POST'])
def safety_helmet_detection():
    """
    实时安全帽检测 API
    使用 ModelScope DAMO-YOLO 模型 (iic/cv_tinynas_object-detection_damoyolo_safety-helmet)
    检测安全帽佩戴情况，返回 'safety hat' 和 'no safety hat' 两类结果
    """
    if 'file' not in request.files:
        return jsonify({'error': '没有文件部分'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': '没有选择文件'}), 400
    
    try:
        image_bytes = file.read()
        nparr = np.frombuffer(image_bytes, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        
        if img is None:
            return jsonify({'error': '无法解码图片'}), 400
        
        img_height, img_width = img.shape[:2]
        
        # 保存临时文件供 ModelScope pipeline 使用
        temp_dir = tempfile.gettempdir()
        temp_path = os.path.join(temp_dir, f'safety_helmet_{uuid.uuid4().hex}.jpg')
        cv2.imwrite(temp_path, img)
        
        try:
            pipeline = _get_safety_helmet_pipeline()
            
            if pipeline is not None:
                result = pipeline(temp_path)
                
                scores = result.get('scores', [])
                labels = result.get('labels', [])
                boxes = result.get('boxes', [])
                
                detections = []
                for i, box in enumerate(boxes):
                    if len(box) >= 4:
                        x1, y1, x2, y2 = box[:4]
                        confidence = float(scores[i]) if i < len(scores) else 0.0
                        label = str(labels[i]) if i < len(labels) else 'unknown'
                        
                        x1 = max(0, int(x1))
                        y1 = max(0, int(y1))
                        x2 = min(img_width, int(x2))
                        y2 = min(img_height, int(y2))
                        
                        detections.append({
                            'bbox': {
                                'x1': float(x1),
                                'y1': float(y1),
                                'x2': float(x2),
                                'y2': float(y2),
                                'width': float(x2 - x1),
                                'height': float(y2 - y1)
                            },
                            'label': label,
                            'confidence': round(confidence, 4)
                        })
                
                detector_name = 'ModelScope DAMO-YOLO (Safety Helmet)'
            else:
                return jsonify({
                    'error': '安全帽检测模型未加载',
                    'suggestion': 'ModelScope DAMO-YOLO 模型初始化失败。请检查: '
                                 '1) 服务器网络是否可访问 modelscope.cn; '
                                 '2) 是否在 Docker 构建时预下载了模型; '
                                 '3) 可尝试 docker build --network=host 重新构建。'
                }), 500
        finally:
            try:
                os.remove(temp_path)
            except:
                pass
        
        # 在图片上绘制检测框
        img_draw = img.copy()
        for det in detections:
            bbox = det['bbox']
            x1 = int(bbox['x1'])
            y1 = int(bbox['y1'])
            x2 = int(bbox['x2'])
            y2 = int(bbox['y2'])
            
            # 根据标签选择颜色：安全帽=绿色，未戴安全帽=红色
            if det['label'] == 'safety hat':
                color = (0, 255, 0)  # 绿色
                label_text = f"Safety Hat {det['confidence']:.2f}"
            else:
                color = (0, 0, 255)  # 红色
                label_text = f"No Hat {det['confidence']:.2f}"
            
            cv2.rectangle(img_draw, (x1, y1), (x2, y2), color, 2)
            
            (label_w, label_h), _ = cv2.getTextSize(label_text, cv2.FONT_HERSHEY_SIMPLEX, 0.6, 2)
            cv2.rectangle(img_draw, (x1, y1 - label_h - 10), (x1 + label_w + 10, y1), color, -1)
            cv2.putText(img_draw, label_text, (x1 + 5, y1 - 5),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)
        
        _, buffer_draw = cv2.imencode('.jpg', img_draw, [cv2.IMWRITE_JPEG_QUALITY, 90])
        img_result_base64 = base64.b64encode(buffer_draw).decode('utf-8')
        
        # 统计
        safety_hat_count = sum(1 for d in detections if d['label'] == 'safety hat')
        no_safety_hat_count = sum(1 for d in detections if d['label'] == 'no safety hat')
        
        return jsonify({
            'success': True,
            'detection_count': len(detections),
            'safety_hat_count': safety_hat_count,
            'no_safety_hat_count': no_safety_hat_count,
            'detections': detections,
            'image_width': img_width,
            'image_height': img_height,
            'result_image': f'data:image/jpeg;base64,{img_result_base64}',
            'detector': detector_name,
            'model_ref': 'https://modelscope.cn/models/iic/cv_tinynas_object-detection_damoyolo_safety-helmet'
        })
        
    except Exception as e:
        return jsonify({'error': f'安全帽检测失败: {str(e)}'}), 500

# ============================================================
# 车牌识别 API（基于 PaddleOCR 全图扫描 + 车牌号过滤）
# ============================================================

# 全局缓存 PaddleOCR 实例（用于车牌文字识别）
_license_ocr_lock = threading.Lock()
_license_ocr_instance = None

# 最大重试次数
_LICENSE_OCR_MAX_RETRIES = 3
# 重试间隔（秒）
_LICENSE_OCR_RETRY_DELAY = 5

def _init_license_ocr():
    """初始化 PaddleOCR 实例（内部方法，已持有锁时调用）"""
    global _license_ocr_instance
    try:
        from paddleocr import PaddleOCR
        # 使用 PP-OCRv5 模型，优化车牌识别效果
        _license_ocr_instance = PaddleOCR(
            lang='ch',
            ocr_version='PP-OCRv5',
            text_det_box_thresh=0.13,
            text_det_unclip_ratio=2.5,
            text_rec_score_thresh=0.1,
            use_doc_unwarping=False
        )
        print("[LicensePlate] PaddleOCR 车牌识别模型已加载 (PP-OCRv5)")
        return True
    except Exception as e:
        print(f"[LicensePlate] 加载 PaddleOCR 失败: {e}")
        _license_ocr_instance = None
        return False

def _get_license_ocr():
    """获取或初始化 PaddleOCR 实例（延迟加载，线程安全，带重试机制）"""
    global _license_ocr_instance
    if _license_ocr_instance is None:
        with _license_ocr_lock:
            if _license_ocr_instance is None:
                # 带重试的初始化
                for attempt in range(1, _LICENSE_OCR_MAX_RETRIES + 1):
                    print(f"[LicensePlate] 尝试初始化 PaddleOCR (第 {attempt}/{_LICENSE_OCR_MAX_RETRIES} 次)...")
                    if _init_license_ocr():
                        break
                    if attempt < _LICENSE_OCR_MAX_RETRIES:
                        print(f"[LicensePlate] 等待 {_LICENSE_OCR_RETRY_DELAY} 秒后重试...")
                        import time
                        time.sleep(_LICENSE_OCR_RETRY_DELAY)
                    else:
                        print(f"[LicensePlate] PaddleOCR 初始化失败，已重试 {_LICENSE_OCR_MAX_RETRIES} 次")
    return _license_ocr_instance

def preload_license_ocr():
    """预加载车牌识别模型（在应用启动时调用）"""
    print("[LicensePlate] 正在预加载车牌识别模型...")
    ocr = _get_license_ocr()
    if ocr is not None:
        print("[LicensePlate] 车牌识别模型预加载成功")
        # 执行一次预测以预热模型
        try:
            import numpy as np
            test_img = np.ones((100, 100, 3), dtype=np.uint8) * 255
            ocr.predict(test_img)
            print("[LicensePlate] 车牌识别模型预热完成")
        except Exception as e:
            print(f"[LicensePlate] 模型预热失败（不影响后续使用）: {e}")
    else:
        print("[LicensePlate] 车牌识别模型预加载失败，将在首次请求时重试")

def _is_license_plate_text(text):
    """
    判断文本是否可能是车牌号
    
    中国车牌号规则:
    - 新能源: 6位 (省+字母+数字/字母混合)
    - 普通蓝牌: 7位 (省+字母+5位数字/字母)
    - 包含汉字、字母、数字混合
    - 长度通常为 6-8 个字符
    """
    if not text:
        return False
    
    # 去除空格
    text = text.strip()
    
    # 车牌号通常长度在 6-8 个字符
    if len(text) < 6 or len(text) > 8:
        return False
    
    # 检查是否包含字母和数字的混合
    has_digit = any(c.isdigit() for c in text)
    has_alpha = any(c.isalpha() for c in text)
    
    # 车牌必须包含字母和数字（或纯数字/纯字母的短车牌）
    if not (has_digit or has_alpha):
        return False
    
    # 检查是否包含常见车牌汉字（省份简称）
    province_chars = {'京', '津', '沪', '渝', '冀', '豫', '云', '辽', '黑', '湘',
                      '皖', '鲁', '新', '苏', '浙', '赣', '鄂', '桂', '甘', '晋',
                      '蒙', '陕', '吉', '闽', '贵', '粤', '川', '青', '藏', '琼',
                      '宁', '港', '澳', '台', '使', '领'}
    
    has_province = any(c in province_chars for c in text)
    
    # 如果有省份简称，大概率是车牌
    if has_province:
        return True
    
    # 没有省份简称但长度合适且包含字母数字混合，也可能是车牌（如部分识别不全的情况）
    if has_digit and has_alpha and len(text) >= 6:
        return True
    
    return False

@app.route('/api/license_plate_detection', methods=['POST'])
def license_plate_detection():
    """
    车牌识别 API
    
    使用 PaddleOCR 全图扫描，自动检测并识别车牌号。
    通过车牌号格式规则（长度、字符组成、省份简称等）过滤出车牌结果。
    """
    if 'file' not in request.files:
        return jsonify({'error': '没有文件部分'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': '没有选择文件'}), 400
    
    try:
        image_bytes = file.read()
        nparr = np.frombuffer(image_bytes, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        
        if img is None:
            return jsonify({'error': '无法解码图片'}), 400
        
        img_height, img_width = img.shape[:2]
        
        # 使用 PaddleOCR 全图扫描
        ocr = _get_license_ocr()
        
        detections = []
        
        if ocr is not None:
            try:
                # 使用 PaddleOCR predict() 方法进行全图 OCR 识别
                # 需要将 BGR 转为 RGB
                img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                ocr_result = ocr.predict(img_rgb)
                
                if ocr_result and len(ocr_result) > 0:
                    ocr_data = ocr_result[0]
                    if isinstance(ocr_data, dict):
                        rec_texts = ocr_data.get('rec_texts', [])
                        rec_scores = ocr_data.get('rec_scores', [])
                        rec_polys = ocr_data.get('rec_polys', [])
                        
                        for i in range(len(rec_texts)):
                            text = rec_texts[i]
                            confidence = float(rec_scores[i]) if i < len(rec_scores) else 0.0
                            
                            # 判断是否为车牌号
                            if _is_license_plate_text(text):
                                # 获取多边形坐标
                                poly = rec_polys[i] if i < len(rec_polys) else None
                                
                                if poly is not None and isinstance(poly, np.ndarray):
                                    xs = poly[:, 0]
                                    ys = poly[:, 1]
                                else:
                                    continue
                                
                                x1, y1 = int(min(xs)), int(min(ys))
                                x2, y2 = int(max(xs)), int(max(ys))
                                
                                detections.append({
                                    'bbox': {
                                        'x1': float(x1),
                                        'y1': float(y1),
                                        'x2': float(x2),
                                        'y2': float(y2),
                                        'width': float(x2 - x1),
                                        'height': float(y2 - y1)
                                    },
                                    'confidence': round(confidence, 4),
                                    'plate_text': text,
                                    'plate_text_confidence': round(confidence, 4)
                                })
                
                detector_name = 'PaddleOCR (License Plate Recognition)'
            except Exception as ocr_err:
                error_msg = str(ocr_err)
                print(f"[LicensePlate] OCR 识别失败: {error_msg}")
                # 判断是否是模型下载相关错误
                if 'download' in error_msg.lower() or 'connection' in error_msg.lower() or 'timeout' in error_msg.lower() or 'http' in error_msg.lower():
                    return jsonify({
                        'error': f'OCR模型加载失败: {error_msg}',
                        'suggestion': '请检查服务器网络连接，确保可以访问PaddleOCR模型下载地址。'
                                     '如果使用Docker，请尝试: docker build --network=host 或在Dockerfile中预下载模型。'
                                     '也可以手动下载模型后挂载到 /root/.paddleocr 目录。'
                    }), 500
                return jsonify({'error': f'OCR识别失败: {error_msg}'}), 500
        else:
            return jsonify({
                'error': 'OCR 模型未加载',
                'suggestion': 'PaddleOCR模型初始化失败。请检查: '
                             '1) 服务器网络是否可访问PaddleOCR模型下载地址; '
                             '2) 是否在Docker构建时预下载了模型; '
                             '3) 模型文件是否完整。'
                             '可尝试重启容器或重新构建镜像。'
            }), 500
        
        # 在图片上绘制检测框和识别结果
        # 使用 PIL 绘制中文文本（cv2.putText 不支持中文）
        from PIL import ImageDraw, ImageFont
        
        img_draw = img.copy()
        img_pil = Image.fromarray(cv2.cvtColor(img_draw, cv2.COLOR_BGR2RGB))
        draw = ImageDraw.Draw(img_pil)
        
        # 尝试加载中文字体（优先使用项目自带字体，兼容 Windows/Linux/macOS）
        _script_dir = os.path.dirname(os.path.abspath(__file__))
        _project_root = os.path.dirname(_script_dir)
        font_paths = [
            os.path.join(_project_root, 'frontend', 'static', 'fonts', 'NotoSansSC.ttf'),
            '/usr/share/fonts/truetype/wqy/wqy-microhei.ttc',  # Linux 文泉驿
            '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc',  # Linux Noto
            'C:/Windows/Fonts/msyh.ttc',       # 微软雅黑 (Windows)
            'C:/Windows/Fonts/simhei.ttf',      # 黑体 (Windows)
            'C:/Windows/Fonts/simsun.ttc',      # 宋体 (Windows)
        ]
        chinese_font = None
        for fp in font_paths:
            if os.path.exists(fp):
                chinese_font = ImageFont.truetype(fp, 24)
                break
        if chinese_font is None:
            # fallback: 使用默认字体
            chinese_font = ImageFont.load_default()
        
        for det in detections:
            bbox = det['bbox']
            x1 = int(bbox['x1'])
            y1 = int(bbox['y1'])
            x2 = int(bbox['x2'])
            y2 = int(bbox['y2'])
            
            # 绘制矩形框（蓝色）
            color_bgr = (255, 0, 0)
            color_rgb = (0, 0, 255)  # PIL 使用 RGB
            cv2.rectangle(img_draw, (x1, y1), (x2, y2), color_bgr, 3)
            
            # 绘制标签：车牌号（使用 PIL 支持中文）
            plate_text = det.get('plate_text', '')
            label_text = f"{plate_text}"
            
            # 使用 PIL 绘制中文文本
            bbox_text = draw.textbbox((0, 0), label_text, font=chinese_font)
            label_w = bbox_text[2] - bbox_text[0]
            label_h = bbox_text[3] - bbox_text[1]
            
            # 标签背景（用 cv2 绘制）
            cv2.rectangle(img_draw, (x1, y1 - label_h - 12), (x1 + label_w + 12, y1), color_bgr, -1)
            
            # 用 PIL 绘制中文文本在 cv2 图像上
            img_pil = Image.fromarray(cv2.cvtColor(img_draw, cv2.COLOR_BGR2RGB))
            draw = ImageDraw.Draw(img_pil)
            draw.text((x1 + 6, y1 - label_h - 6), label_text, font=chinese_font, fill=(255, 255, 255))
            img_draw = cv2.cvtColor(np.array(img_pil), cv2.COLOR_RGB2BGR)
            
            # 在框下方显示置信度（纯英文数字，cv2 可以正常显示）
            conf_text = f"conf: {det['confidence']:.2f}"
            cv2.putText(img_draw, conf_text, (x1, y2 + 20),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.5, color_bgr, 2)
        
        _, buffer_draw = cv2.imencode('.jpg', img_draw, [cv2.IMWRITE_JPEG_QUALITY, 90])
        img_result_base64 = base64.b64encode(buffer_draw).decode('utf-8')
        
        return jsonify({
            'success': True,
            'detection_count': len(detections),
            'detections': detections,
            'image_width': img_width,
            'image_height': img_height,
            'result_image': f'data:image/jpeg;base64,{img_result_base64}',
            'detector': detector_name,
            'model_ref': 'https://modelscope.cn/models/iic/cv_resnet18_license-plate-detection_damo'
        })
        
    except Exception as e:
        return jsonify({'error': f'车牌检测失败: {str(e)}'}), 500

# ============================================================
# 全身关键点检测 API
# ============================================================

# 全局缓存 Keypoint R-CNN 模型（线程安全）
_keypoint_model_lock = threading.Lock()
_keypoint_model = None
_keypoint_device = None

# COCO 关键点名称 (17个关键点)
KEYPOINT_NAMES = [
    'nose', 'left_eye', 'right_eye', 'left_ear', 'right_ear',
    'left_shoulder', 'right_shoulder', 'left_elbow', 'right_elbow',
    'left_wrist', 'right_wrist', 'left_hip', 'right_hip',
    'left_knee', 'right_knee', 'left_ankle', 'right_ankle'
]

# 骨架连接
KEYPOINT_SKELETON = [
    (0, 1), (0, 2), (1, 3), (2, 4),
    (5, 6), (5, 7), (7, 9), (6, 8), (8, 10),
    (5, 11), (6, 12), (11, 12),
    (11, 13), (12, 14), (13, 15), (14, 16)
]

# 关键点颜色
KEYPOINT_COLORS = [
    (0, 0, 255), (255, 0, 0), (0, 255, 0), (0, 255, 255), (255, 0, 255),
    (255, 255, 0), (128, 0, 255), (0, 128, 255), (255, 128, 0),
    (0, 255, 128), (128, 255, 0), (255, 0, 128), (128, 0, 0),
    (0, 0, 128), (0, 128, 0), (128, 128, 0), (0, 128, 128)
]

# 最大重试次数
_KEYPOINT_MAX_RETRIES = 3
_KEYPOINT_RETRY_DELAY = 5

def _init_keypoint_model():
    """初始化 Keypoint R-CNN 模型（内部方法，已持有锁时调用）"""
    global _keypoint_model, _keypoint_device
    try:
        import torch
        import torchvision
        _keypoint_model = torchvision.models.detection.keypointrcnn_resnet50_fpn(
            weights=torchvision.models.detection.KeypointRCNN_ResNet50_FPN_Weights.DEFAULT
        )
        _keypoint_model.eval()
        _keypoint_device = torch.device('cuda') if torch.cuda.is_available() else torch.device('cpu')
        _keypoint_model.to(_keypoint_device)
        print(f"[KeypointDetection] Keypoint R-CNN 模型已加载 (设备: {_keypoint_device})")
        return True
    except Exception as e:
        print(f"[KeypointDetection] 加载 Keypoint R-CNN 模型失败: {e}")
        _keypoint_model = None
        return False

def _get_keypoint_model():
    """获取或初始化 Keypoint R-CNN 模型（延迟加载，线程安全，带重试）"""
    global _keypoint_model, _keypoint_device
    if _keypoint_model is None:
        with _keypoint_model_lock:
            if _keypoint_model is None:
                for attempt in range(1, _KEYPOINT_MAX_RETRIES + 1):
                    print(f"[KeypointDetection] 尝试初始化 (第 {attempt}/{_KEYPOINT_MAX_RETRIES} 次)...")
                    if _init_keypoint_model():
                        break
                    if attempt < _KEYPOINT_MAX_RETRIES:
                        print(f"[KeypointDetection] 等待 {_KEYPOINT_RETRY_DELAY} 秒后重试...")
                        import time
                        time.sleep(_KEYPOINT_RETRY_DELAY)
                    else:
                        print(f"[KeypointDetection] 初始化失败，已重试 {_KEYPOINT_MAX_RETRIES} 次")
    return _keypoint_model, _keypoint_device

def preload_keypoint():
    """预加载关键点检测模型（在应用启动时调用）"""
    print("[KeypointDetection] 正在预加载关键点检测模型...")
    model, device = _get_keypoint_model()
    if model is not None:
        print(f"[KeypointDetection] 关键点检测模型预加载成功 (设备: {device})")
        # 预热模型
        try:
            import torch
            import numpy as np
            test_img = np.ones((100, 100, 3), dtype=np.uint8) * 128
            img_tensor = torch.from_numpy(test_img).permute(2, 0, 1).float().div(255.0).unsqueeze(0).to(device)
            with torch.no_grad():
                model(img_tensor)
            print("[KeypointDetection] 关键点检测模型预热完成")
        except Exception as e:
            print(f"[KeypointDetection] 模型预热失败（不影响后续使用）: {e}")
    else:
        print("[KeypointDetection] 关键点检测模型预加载失败，将在首次请求时重试")

@app.route('/api/keypoint_detection', methods=['POST'])
def keypoint_detection():
    """
    全身关键点检测 API
    使用 PyTorch Keypoint R-CNN (ResNet50-FPN) 模型
    检测 17 个 COCO 人体关键点
    参考模型: iic/cv_hrnetw48_human-wholebody-keypoint_image (ModelScope)
    """
    if 'file' not in request.files:
        return jsonify({'error': '没有文件部分'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': '没有选择文件'}), 400
    
    try:
        image_bytes = file.read()
        nparr = np.frombuffer(image_bytes, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        
        if img is None:
            return jsonify({'error': '无法解码图片'}), 400
        
        img_height, img_width = img.shape[:2]
        
        # 使用 Keypoint R-CNN 模型
        model, device = _get_keypoint_model()
        
        if model is not None:
            import torch
            # 转换为 RGB (torchvision 需要 RGB)
            img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            img_tensor = torch.from_numpy(img_rgb).permute(2, 0, 1).float().div(255.0).unsqueeze(0).to(device)
            
            with torch.no_grad():
                predictions = model(img_tensor)
            
            scores = predictions[0]['scores'].cpu().numpy()
            boxes = predictions[0]['boxes'].cpu().numpy()
            keypoints = predictions[0]['keypoints'].cpu().numpy()
            
            # 过滤低置信度检测
            confidence_threshold = 0.5
            valid_indices = scores > confidence_threshold
            
            persons = []
            for idx in range(len(scores)):
                if not valid_indices[idx]:
                    continue
                
                score = float(scores[idx])
                box = boxes[idx]
                kps = keypoints[idx]
                
                x1, y1, x2, y2 = map(int, box)
                x1 = max(0, x1); y1 = max(0, y1)
                x2 = min(img_width, x2); y2 = min(img_height, y2)
                
                person_kps = []
                for j, kp in enumerate(kps):
                    x, y, conf = int(kp[0]), int(kp[1]), float(kp[2])
                    person_kps.append({
                        'name': KEYPOINT_NAMES[j] if j < len(KEYPOINT_NAMES) else f'kp_{j}',
                        'x': x, 'y': y,
                        'confidence': round(conf, 4)
                    })
                
                persons.append({
                    'bbox': {
                        'x1': float(x1), 'y1': float(y1),
                        'x2': float(x2), 'y2': float(y2),
                        'width': float(x2 - x1), 'height': float(y2 - y1)
                    },
                    'confidence': round(score, 4),
                    'keypoints': person_kps
                })
            
            detector_name = 'PyTorch Keypoint R-CNN (ResNet50-FPN)'
        else:
            return jsonify({
                'error': '关键点检测模型未加载',
                'suggestion': 'PyTorch Keypoint R-CNN 模型初始化失败。请检查: '
                             '1) 服务器网络是否可访问 PyTorch 模型下载地址; '
                             '2) 是否在 Docker 构建时预下载了模型; '
                             '3) 可尝试 docker build --network=host 重新构建。'
            }), 500
        
        # 在图片上绘制检测结果
        img_draw = img.copy()
        for person in persons:
            bbox = person['bbox']
            x1, y1, x2, y2 = int(bbox['x1']), int(bbox['y1']), int(bbox['x2']), int(bbox['y2'])
            
            # 绘制检测框
            cv2.rectangle(img_draw, (x1, y1), (x2, y2), (0, 255, 0), 2)
            cv2.putText(img_draw, f"Person {person['confidence']:.2f}", (x1, y1-10),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
            
            # 绘制关键点
            for j, kp in enumerate(person['keypoints']):
                if kp['confidence'] > 0.3:
                    color = KEYPOINT_COLORS[j % len(KEYPOINT_COLORS)]
                    cv2.circle(img_draw, (kp['x'], kp['y']), 4, color, -1)
                    cv2.circle(img_draw, (kp['x'], kp['y']), 5, (255, 255, 255), 1)
            
            # 绘制骨架
            for (start, end) in KEYPOINT_SKELETON:
                kps_list = person['keypoints']
                if start < len(kps_list) and end < len(kps_list):
                    if kps_list[start]['confidence'] > 0.3 and kps_list[end]['confidence'] > 0.3:
                        pt1 = (kps_list[start]['x'], kps_list[start]['y'])
                        pt2 = (kps_list[end]['x'], kps_list[end]['y'])
                        cv2.line(img_draw, pt1, pt2, (0, 255, 255), 2)
        
        # 编码结果图片
        _, buffer_draw = cv2.imencode('.jpg', img_draw, [cv2.IMWRITE_JPEG_QUALITY, 90])
        img_result_base64 = base64.b64encode(buffer_draw).decode('utf-8')
        
        return jsonify({
            'success': True,
            'person_count': len(persons),
            'persons': persons,
            'image_width': img_width,
            'image_height': img_height,
            'result_image': f'data:image/jpeg;base64,{img_result_base64}',
            'detector': detector_name,
            'model_ref': 'https://modelscope.cn/models/iic/cv_hrnetw48_human-wholebody-keypoint_image'
        })
        
    except Exception as e:
        return jsonify({'error': f'关键点检测失败: {str(e)}'}), 500

if __name__ == '__main__':
    print("启动OCR工业图片识别系统...")
    print(f"上传目录: {app.config['UPLOAD_FOLDER']}")
    print(f"调试模式: {'开启' if Config.DEBUG else '关闭'}")
    print(f"CORS支持: {'开启' if Config.ENABLE_CORS else '关闭'}")
    
    # 显示OCR引擎信息
    engine_info = ocr_processor.get_engine_info()
    print(f"OCR引擎: {engine_info['current_engine']} (类型: {engine_info['engine_type']})")
    print(f"PaddleOCR可用: {engine_info['paddleocr_available']}")
    print(f"OpenAI VL可用: {engine_info['openai_vl_available']}")
    
    # 预加载所有 AI 模型
    print("\n" + "=" * 60)
    print("[启动预加载] 正在预加载各功能模型...")
    print("=" * 60)
    
    # 1. 车牌识别模型 (PaddleOCR PP-OCRv5)
    preload_license_ocr()
    
    # 2. 安全帽检测模型 (ModelScope DAMO-YOLO)
    preload_safety_helmet()
    
    # 3. 关键点检测模型 (PyTorch Keypoint R-CNN)
    preload_keypoint()
    
    print("=" * 60)
    print("[启动预加载] 所有模型预加载完成")
    print("=" * 60 + "\n")
    
    print(f"API服务运行在 http://127.0.0.1:{Config.PORT}")
    
    # 验证配置
    config_issues = Config.validate_config()
    if config_issues:
        print("配置警告:")
        for issue in config_issues:
            print(f"  - {issue}")
    
    app.run(debug=Config.DEBUG, host='0.0.0.0', port=Config.PORT)